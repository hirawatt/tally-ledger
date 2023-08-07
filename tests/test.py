from openpyxl import load_workbook
from yattag import Doc, indent
  
# Load our Excel File
wb = load_workbook(".\data\Accounting Voucher.xlsx")
# Getting an object of active sheet 1
ws = wb.worksheets[0]
  
# Returning returns a triplet
doc, tag, text = Doc().tagtext()
  
xml_header = '<?xml version="1.0" encoding="UTF-8"?>'
xml_schema = '<xs:schema xmlns:xs="http://www.w3.org/2001/XMLSchema"></xs:schema>'
  
# Appends the String to document
doc.asis(xml_header)
doc.asis(xml_schema)
  
with tag('People'):
    for row in ws.iter_rows(min_row=2, max_row=10, min_col=1, max_col=6):
        row = [cell.value for cell in row]
        with tag("Person"):
            with tag("First_Name"):
                text(row[0])
            with tag("Last_Name"):
                text(row[1])
            with tag("Gender"):
                text(row[2])
            with tag("Country"):
                text(row[3])
            with tag("Age"):
                text(row[4])
            with tag("Date"):
                text(row[5])
  
result = indent(
    doc.getvalue(),
    indentation='   ',
    indent_text=True
)
  
with open("output.xml", "w") as f:
    f.write(result)